from django.http import HttpResponse
from openpyxl import Workbook

from CLIENTES.models import Cliente


def get(self,request,args,**kwargs):
    clientes = Cliente.objects.all()
    wb = Workbook()
    ws = wb.active
    ws['B1'] = "REPORTE DE AUTORES"

    ws.marge_cells('B1:E1')
    ws['B3'] = "NOMBRES"
    ws['C3'] = "APELLIDOS"
    ws['D3'] = "DIRECCION"
    ws['E1'] = "EMAIL"

    cont = 4

    for cliente in clientes:
        ws.cells(row = cont, column = 2).value = cliente.nombre
        ws.cells(row=cont, column=2).value = cliente.apellido
        ws.cells(row=cont, column=2).value = cliente.direccion
        ws.cells(row=cont, column=2).value = cliente.email
        cont+=1

    nombre_archivo = "ReporteClienteExcel.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition']= content
    wb.save(response)
    return response
